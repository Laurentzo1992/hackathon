from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from .models import Hackaton, HackathonUser, Inscription, Programme, Partenaire
from django.urls import path
from django.utils.html import format_html
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Spacer


class HackathonUserAdmin(UserAdmin):
    model = HackathonUser
    list_display = ("last_name", "first_name", "email", "phone", "is_staff", "is_active")
    search_fields = ("email", "phone")
    ordering = ("email",)

    fieldsets = (
        (None, {"fields": ("email", "password", "phone", "address", "birth_date")}),
        ("Permissions", {"fields": ("is_staff", "is_active", "is_superuser", "groups", "user_permissions")}),
        ("Dates importantes", {"fields": ("last_login", "date_joined")}),
    )

    add_fieldsets = (
        (None, {
            "classes": ("wide",),
            "fields": ("email", "password1", "password2", "phone", "address", "birth_date", "is_staff", "is_active")}
        ),
    )
    
    



@admin.register(Hackaton)
class HackatonAdmin(admin.ModelAdmin):
    list_display = ("theme", "dateOuverture", "dateFermeture", "statut", "export_links")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("<int:hackaton_id>/export_xlsx/", self.admin_site.admin_view(self.export_xlsx), name="hackaton_export_xlsx"),
            path("<int:hackaton_id>/export_pdf/", self.admin_site.admin_view(self.export_pdf), name="hackaton_export_pdf"),
        ]
        return custom_urls + urls

    def export_links(self, obj):
        return format_html(
            '<a class="button" href="{}">📊 Excel</a> &nbsp; <a class="button" href="{}">📄 PDF</a>',
            f"{obj.id}/export_xlsx/",
            f"{obj.id}/export_pdf/",
        )
    export_links.short_description = "Exports"

    def export_xlsx(self, request, hackaton_id):
        hackaton = Hackaton.objects.get(id=hackaton_id)
        inscriptions = Inscription.objects.filter(hackaton=hackaton)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inscriptions"

        headers = ["Nom complet", "Email", "Téléphone", "Niveau d'étude",
                "Situation professionnelle", "Compétence principale", "Équipe", 
                "Projet", "CV", "Dossier de projet"]
        ws.append(headers)

        for inscription in inscriptions:
            row = [
                inscription.nomComplet,
                inscription.user.email if inscription.user else "",
                inscription.user.phone if inscription.user else "",
                inscription.niveauEtude,
                inscription.situationProfessionnelle,
                inscription.competencePrincipale,
                inscription.nomEquipe,
                inscription.ideeProjet,
            ]

            # Ajouter liens si fichiers présents
            cv_link = ""
            piece_link = ""
            if inscription.user and inscription.cv:
                cv_url = request.build_absolute_uri(inscription.cv.url)
                cv_link = "CV"
            if inscription.user and inscription.dossier_projet:
                piece_url = request.build_absolute_uri(inscription.dossier_projet.url)
                piece_link = "Dossier"

            row.extend([cv_link, piece_link])
            ws.append(row)

            # Hyperliens
            if cv_link:
                ws.cell(row=ws.max_row, column=9).hyperlink = cv_url
                ws.cell(row=ws.max_row, column=9).style = "Hyperlink"
            if piece_link:
                ws.cell(row=ws.max_row, column=10).hyperlink = piece_url
                ws.cell(row=ws.max_row, column=10).style = "Hyperlink"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="inscriptions_{hackaton.theme}.xlsx"'
        wb.save(response)
        return response




 



    def export_pdf(self, request, hackaton_id):
        hackaton = Hackaton.objects.get(id=hackaton_id)
        inscriptions = Inscription.objects.filter(hackaton=hackaton)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="inscriptions_{hackaton.theme}.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            leftMargin=20,
            rightMargin=20,
            topMargin=30,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()
        styleN = styles['Normal']

        elements = []

        # Entête du document
        elements.append(Paragraph(f"Liste des inscrits - {hackaton.theme}", styles['Title']))
        elements.append(Spacer(1, 12))

        # Entêtes du tableau
        data = [[
            "Nom", "Email", "Tél", "Niveau",
            "Sit pro", "Compétence",
            "Nom équipe", "Projet", "CV", "Dossier"
        ]]

        # Ajouter les inscriptions
        for ins in inscriptions:
            email = ins.user.email if ins.user else ""
            phone = ins.user.phone if ins.user else ""

            # Texte et liens
            if ins.cv:
                cv_paragraph = Paragraph(f'<link href="{request.build_absolute_uri(ins.cv.url)}">CV</link>', styleN)
            else:
                cv_paragraph = Paragraph("-", styleN)

            if ins.dossier_projet:
                dossier_paragraph = Paragraph(f'<link href="{request.build_absolute_uri(ins.dossier_projet.url)}">Dossier</link>', styleN)
            else:
                dossier_paragraph = Paragraph("-", styleN)

            row = [
                Paragraph(ins.nomComplet or "", styleN),
                Paragraph(email, styleN),
                Paragraph(phone, styleN),
                Paragraph(ins.niveauEtude or "", styleN),
                Paragraph(ins.situationProfessionnelle or "", styleN),
                Paragraph(ins.competencePrincipale or "", styleN),
                Paragraph(ins.nomEquipe or "", styleN),
                Paragraph(ins.ideeProjet or "", styleN),
                cv_paragraph,
                dossier_paragraph
            ]
            data.append(row)

        # Largeurs des colonnes pour éviter débordement
        col_widths = [50, 50, 30, 40, 60, 60, 80, 70, 30, 45]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#d3d3d3')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))

        elements.append(table)
        doc.build(elements)
        return response




admin.site.register(HackathonUser, HackathonUserAdmin)
admin.site.register(Programme)
admin.site.register(Partenaire)
